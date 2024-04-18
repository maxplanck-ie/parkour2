import itertools
import json

from common.utils import get_date_range
from common.views import CsrfExemptSessionAuthentication
from django.apps import apps
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .serializers import RunsSerializer, SequencesSerializer

Request = apps.get_model("request", "Request")
Library = apps.get_model("library", "Library")
Sample = apps.get_model("sample", "Sample")
Flowcell = apps.get_model("flowcell", "Flowcell")
Lane = apps.get_model("flowcell", "Lane")


class RunStatisticsViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = RunsSerializer

    def get_queryset(self):
        request_qs = Request.objects.filter(archived=False).only("name")

        libraries_qs = (
            Library.objects.filter(~Q(status=-1))
            .select_related(
                "read_length",
                "library_protocol",
                "library_type",
            )
            .prefetch_related(
                Prefetch("request", queryset=request_qs, to_attr="fetched_request")
            )
            .only(
                "read_length__name",
                "library_protocol__name",
                "library_type__name",
            )
        )

        samples_qs = (
            Sample.objects.filter(~Q(status=-1))
            .select_related(
                "read_length",
                "library_protocol",
                "library_type",
            )
            .prefetch_related(
                Prefetch("request", queryset=request_qs, to_attr="fetched_request")
            )
            .only(
                "read_length__name",
                "library_protocol__name",
                "library_type__name",
            )
        )

        lanes_qs = (
            Lane.objects.select_related("pool")
            .prefetch_related(
                Prefetch(
                    "pool__libraries",
                    queryset=libraries_qs,
                    to_attr="fetched_libraries",
                ),
                Prefetch(
                    "pool__samples",
                    queryset=samples_qs,
                    to_attr="fetched_samples",
                ),
            )
            .only(
                "name",
                "phix",
                "loading_concentration",
                "pool__name",
                "pool__libraries",
                "pool__samples",
            )
        )

        queryset = (
            Flowcell.objects.filter(archived=False)
            .exclude(
                matrix__isnull=True,
            )
            .select_related(
                "pool_size",
            )
            .prefetch_related(
                Prefetch("lanes", queryset=lanes_qs, to_attr="fetched_lanes"),
            )
            .order_by("-create_time")
        )

        return queryset

    def list(self, request):
        now = timezone.now()
        start = request.query_params.get("start", now)
        end = request.query_params.get("end", now)
        start, end = get_date_range(start, end, "%Y-%m-%dT%H:%M:%S")

        queryset = self.filter_queryset(self.get_queryset()).filter(
            create_time__gte=start,
            create_time__lte=end,
        )

        if request.GET.get("asHandler") == "True":
            queryset = queryset.filter(requests__handler=request.user)

        if request.GET.get("asBioinformatician") == "True":
            queryset = queryset.filter(requests__bioinformatician=request.user)

        if request.user.is_staff or request.user.member_of_bcf:
            pass
        elif request.user.is_pi:
            queryset = queryset.filter(requests__pi=request.user)
        else:
            queryset = queryset.filter(Q(requests__user=request.user) | Q(requests__bioinformatician=request.user)).distinct()

        serializer = self.get_serializer(queryset, many=True)
        data = list(itertools.chain(*serializer.data))
        return Response(data)

    @action(methods=["post"], detail=False)
    def upload(self, request):
        flowcell_id = request.data.get("flowcell_id", "")
        matrix = request.data.get("matrix", "")

        try:
            flowcell = Flowcell.objects.filter(archived=False).get(
                flowcell_id=flowcell_id
            )
        except (ValueError, Flowcell.DoesNotExist):
            return Response(
                {
                    "success": False,
                    "message": f'Flowcell with id "{flowcell_id}" doesn\'t exist.',
                },
                400,
            )

        try:
            matrix = json.loads(matrix)
        except ValueError:
            return Response(
                {
                    "success": False,
                    "message": "Invalid matrix data.",
                },
                400,
            )

        # Update pre-existing information
        currentMatrix = list()
        if flowcell.matrix:
            currentMatrix = flowcell.matrix

        found = dict()
        for idx, entry in enumerate(currentMatrix):
            found[entry["name"]] = idx
        for entry in matrix:
            if entry["name"] in found:
                currentMatrix[found[entry["name"]]] = entry
            else:
                currentMatrix.append(entry)

        flowcell.matrix = currentMatrix
        flowcell.save(update_fields=["matrix"])
        return Response({"success": True})


class SequencesStatisticsViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = SequencesSerializer

    def get_queryset(self):
        libraries_qs = (
            Library.objects.filter(~Q(status=-1))
            .select_related(
                "library_protocol",
                "library_type",
            )
            .only(
                "name",
                "barcode",
                "library_protocol__name",
                "library_type__name",
            )
        )

        samples_qs = (
            Sample.objects.filter(~Q(status=-1))
            .select_related(
                "library_protocol",
                "library_type",
            )
            .only(
                "name",
                "barcode",
                "library_protocol__name",
                "library_type__name",
            )
        )

        requests_qs = (
            Request.objects.filter(archived=False)
            .prefetch_related(
                Prefetch(
                    "libraries",
                    queryset=libraries_qs,
                    to_attr="fetched_libraries",
                ),
                Prefetch(
                    "samples",
                    queryset=samples_qs,
                    to_attr="fetched_samples",
                ),
            )
            .only(
                "name",
                "libraries",
                "samples",
            )
        )

        lanes_qs = (
            Lane.objects.select_related("pool")
            .prefetch_related(
                Prefetch(
                    "pool__libraries",
                    queryset=Library.objects.only("barcode"),
                    to_attr="fetched_libraries",
                ),
                Prefetch(
                    "pool__samples",
                    queryset=Sample.objects.only("barcode"),
                    to_attr="fetched_samples",
                ),
            )
            .only(
                "name",
                "pool__name",
                "pool__libraries",
                "pool__samples",
            )
            .order_by("name")
        )

        queryset = (
            Flowcell.objects.filter(archived=False)
            .exclude(
                sequences__isnull=True,
            )
            .select_related(
                "pool_size",
            )
            .prefetch_related(
                Prefetch(
                    "lanes",
                    queryset=lanes_qs,
                    to_attr="fetched_lanes",
                ),
                Prefetch(
                    "requests",
                    queryset=requests_qs,
                    to_attr="fetched_requests",
                ),
            )
            .order_by("-create_time")
        )

        return queryset

    def list(self, request):
        now = timezone.now()
        start = request.query_params.get("start", now)
        end = request.query_params.get("end", now)
        start, end = get_date_range(start, end, "%Y-%m-%dT%H:%M:%S")

        queryset = self.filter_queryset(self.get_queryset()).filter(
            create_time__gte=start,
            create_time__lte=end,
        )

        if request.GET.get("asHandler") == "True":
            queryset = queryset.filter(requests__handler=request.user)

        if request.GET.get("asBioinformatician") == "True":
            queryset = queryset.filter(requests__bioinformatician=request.user)

        if request.user.is_staff or request.user.member_of_bcf:
            pass
        elif request.user.is_pi:
            queryset = queryset.filter(requests__pi=request.user)
        else:
            queryset = queryset.filter(Q(requests__user=request.user) | Q(requests__bioinformatician=request.user)).distinct()

        serializer = self.get_serializer(queryset, many=True)
        data = list(itertools.chain(*serializer.data))
        return Response(data)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        merge_lanes = self.request.GET.get('merge_lanes',
                                           self.request.POST.get(
                                            'merge_lanes', 'True')) == 'True'
        context.update({'merge_lanes': merge_lanes})
        return context

    @action(methods=["post"], detail=False)
    def upload(self, request):
        flowcell_id = request.data.get("flowcell_id", "")
        sequences = request.data.get("sequences", "")
        flowcell = get_object_or_404(Flowcell, flowcell_id=flowcell_id)

        try:
            sequences = json.loads(sequences)
        except ValueError:
            return Response(
                {
                    "success": False,
                    "message": "Invalid sequences data.",
                },
                400,
            )

        # NZ, disable updating pre-existing sequence information,
        # instead always update entire demultiplexing summary anew

        # # Update pre-existing sequence information
        # currentSequences = list()
        # if flowcell.sequences:
        #     currentSequences = flowcell.sequences

        # found = dict()
        # for idx, entry in enumerate(currentSequences):
        #     found[entry["barcode"]] = idx
        # barcodes = []
        # for entry in sequences:
        #     barcodes.append(entry["barcode"])
        #     if entry["barcode"] in found:
        #         currentSequences[found[entry["barcode"]]] = entry
        #     else:
        #         currentSequences.append(entry)

        # flowcell.sequences = currentSequences

        flowcell.sequences = sequences
        flowcell.save(update_fields=["sequences"])

        # If sample/library in demux report, set status to Data delivered,
        # if not, to Data not delivered
        barcodes = [entry["barcode"] for entry in sequences if entry.get("barcode")]
        # Data delivered
        Library.objects.filter(request__flowcell=flowcell,
                               status__gte=5,
                               barcode__in=barcodes).update(status=6)
        Sample.objects.filter(request__flowcell=flowcell,
                              status__gte=5,
                              barcode__in=barcodes).update(status=6)
        # Data not delivered
        Library.objects.filter(request__flowcell=flowcell,
                               status__gte=5). \
                        exclude(barcode__in=barcodes).update(status=7)
        Sample.objects.filter(request__flowcell=flowcell,
                              status__gte=5). \
                       exclude(request__flowcell=flowcell,
                               barcode__in=barcodes).update(status=7)

        return Response({"success": True})

    @action(
        methods=["post"],
        detail=False,
        authentication_classes=[CsrfExemptSessionAuthentication],
    )
    def download_report(self, request):
        barcodes = json.loads(request.data.get("barcodes", "[]"))
        barcodes_map = {b: True for b in barcodes}
        flowcell_ids = json.loads(request.data.get("flowcell_ids", "[]"))

        filename = "Sequences_Statistics_Report.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        queryset = self.filter_queryset(self.get_queryset()).filter(flowcell_id__in=flowcell_ids)
        serializer = self.get_serializer(queryset, many=True)
        data = list(itertools.chain(*serializer.data))

        bold_font = Font(bold=True)
        wb = Workbook()
        ws = wb.active

        header = [
            "Request",
            "Barcode",
            "Name",
            "Lane",
            "Pool",
            "Library Protocol",
            "Library Type",
            "M Reads PF, requested",
            "M Reads PF, sequenced",
        ]
        ws.append(header)
        
        # Make header bold
        [setattr(cell, 'font', bold_font)
         for cell in ws[f'{ws._current_row}:{ws._current_row}']]
        # Set column width to 20
        for col in range(ws.min_column, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        for item in data:
            if item["barcode"] in barcodes_map:
                reads_pf_sequenced = item.get("reads_pf_sequenced", "")
                if reads_pf_sequenced != "":
                    reads_pf_sequenced = round(int(reads_pf_sequenced) / 1_000_000, 1)
                row = [
                    item["request"],
                    item["barcode"],
                    item["name"],
                    item["lane"],
                    item["pool"],
                    item["library_protocol"],
                    item["library_type"],
                    item.get("reads_pf_requested", ""),
                    reads_pf_sequenced,
                ]
                ws.append(row)

        wb.save(response)
        return response
