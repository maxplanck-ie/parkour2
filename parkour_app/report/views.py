from collections import Counter, OrderedDict

import numpy as np
from django.apps import apps
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import Prefetch
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from pandas import DataFrame

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from index_generator.models import PoolSize
from common.models import User

from .sql import LIBRARY_SELECT, QUERY, SAMPLE_JOINS, SAMPLE_SELECT

Organization = apps.get_model("common", "Organization")
# PrincipalInvestigator = apps.get_model("common", "PrincipalInvestigator")
LibraryProtocol = apps.get_model("library_sample_shared", "LibraryProtocol")

Library = apps.get_model("library", "Library")
Sample = apps.get_model("sample", "Sample")
Request = apps.get_model("request", "Request")
Pool = apps.get_model("index_generator", "Pool")
Sequencer = apps.get_model("flowcell", "Sequencer")
Flowcell = apps.get_model("flowcell", "Flowcell")
Lane = apps.get_model("flowcell", "Lane")


class Report:
    def __init__(self, start, end, sequenced=False):

        # Filter libraries and samples by when a request was submitted
        libraries_qs = (
            Library.objects.select_related("library_protocol", "library_type")
            .filter(request__samples_submitted_time__gte=start,
                    request__samples_submitted_time__lte=end)
            .only("id", "library_protocol__name", "library_type__name")
        )

        samples_qs = (
            Sample.objects.select_related("library_protocol", "library_type")
            .filter(request__samples_submitted_time__gte=start,
                    request__samples_submitted_time__lte=end)
            .only("id", "library_protocol__name", "library_type__name")
        )

        # If necessary, report only records from requests that are already sequenced
        if sequenced:
            libraries_qs = libraries_qs.filter(request__sequenced=True)
            samples_qs = samples_qs.filter(request__sequenced=True)

        self.requests = (
            Request.objects.select_related(
                "cost_unit__organization",
                "pi",
            )
            .prefetch_related(
                Prefetch(
                    "libraries",
                    queryset=libraries_qs,
                    to_attr="fetched_libraries"
                ),
                Prefetch("samples",
                         queryset=samples_qs,
                         to_attr="fetched_samples"),
            )
            .filter(samples_submitted_time__gte=start,
                    samples_submitted_time__lte=end)
            .only(
                "id",
                "libraries",
                "samples",
                "cost_unit__organization__name",
                "pi",
            )
        )

        lanes_qs = (
            Lane.objects.select_related("pool")
            .prefetch_related(
                Prefetch(
                    "pool__libraries",
                    queryset=libraries_qs,
                    to_attr="fetched_libraries",
                ),
                Prefetch(
                    "pool__samples",
                    queryset=samples_qs,
                    to_attr="fetched_samples"
                ),
            )
            .only("id",
                  "pool__libraries",
                  "pool__samples")
        )

        self.flowcells = (
            Flowcell.objects.select_related(
                "pool_size__sequencer",
            )
            .prefetch_related(
                Prefetch("lanes",
                         queryset=lanes_qs,
                         to_attr="fetched_lanes"),
            )
            .filter(requests__samples_submitted_time__gte=start,
                    requests__samples_submitted_time__lte=end)
            .only("id",
                  "pool_size__sequencer__name",
                  "lanes")
        )

        if sequenced:
            self.requests = self.requests.filter(sequenced=True)
            self.flowcells = self.flowcells.filter(requests__sequenced=True)

    def get_total_counts(self):
        data = []
        num_requests = self.requests.count()
        num_libraries = 0
        num_samples = 0
        num_libraries_failed = 0
        num_libraries_compromised = 0
        num_samples_failed = 0
        num_samples_compromised = 0

        for req in self.requests:
            fetched_libraries = req.fetched_libraries
            num_libraries += len(fetched_libraries)
            num_libraries_failed += len([l for l in fetched_libraries if l.status == -1])
            num_libraries_compromised += len([l for l in fetched_libraries if l.status == -2])
            fetched_samples = req.fetched_samples
            num_samples += len(fetched_samples)
            num_samples_failed += len([s for s in fetched_samples if s.status == -1])
            num_samples_compromised += len([s for s in fetched_samples if s.status == -2])

        data.append({"type": "Requests",
                     'count': num_requests,
                     'count_failed': None,
                     'count_compromised': None
                     })

        data.append({"type": "Samples",
                     "count": num_samples,
                     'count_failed': num_samples_failed,
                     'count_compromised': num_samples_compromised})

        data.append({"type": "Libraries",
                     "count": num_libraries,
                     'count_failed': num_libraries_failed,
                     'count_compromised': num_libraries_compromised})

        data.append({"type": "Samples + Libraries",
                     "count": num_samples + num_libraries,
                     'count_failed': num_samples_failed + num_libraries_failed,
                     'count_compromised': num_samples_compromised + num_libraries_compromised})

        return data

    def get_organization_counts(self):
        counts = {}

        for req in self.requests:
            organization = req.cost_unit.organization
            org_name = organization.name if organization else "None"
            if org_name not in counts.keys():
                counts[org_name] = {"libraries": 0, "samples": 0, "total": 0, "requests": 0}
            counts[org_name]["libraries"] += len(req.fetched_libraries)
            counts[org_name]["samples"] += len(req.fetched_samples)
            counts[org_name]["total"] += len(req.fetched_libraries) + len(req.fetched_samples)
            counts[org_name]["requests"] += 1

        return self._get_data(counts)

    def get_library_protocol_counts(self):
        counts = {}

        for req in self.requests:
            # Extract Library Protocols
            library_protocols = [x.library_protocol.name for x in req.fetched_libraries]
            sample_protocols = [x.library_protocol.name for x in req.fetched_samples]

            # Merge the counts
            library_cnt = {
                x[0]: {"libraries": x[1]} for x in Counter(library_protocols).items()
            }
            sample_cnt = {
                x[0]: {"samples": x[1]} for x in Counter(sample_protocols).items()
            }
            count = {
                k: {
                    **library_cnt.get(k, {"libraries": 0}),
                    **sample_cnt.get(k, {"samples": 0}),
                }
                for k in library_cnt.keys() | sample_cnt.keys()
            }

            for k, v in count.items():
                temp_dict = counts.get(k, {"libraries": 0, "samples": 0, "total": 0, "requests": 0})
                temp_dict["libraries"] += v["libraries"]
                temp_dict["samples"] += v["samples"]
                temp_dict["total"] += v["libraries"] + v["samples"]
                temp_dict["requests"] += 1
                counts[k] = temp_dict

        return self._get_data(counts)

    def get_library_type_counts(self):

        counts = {}
        for req in self.requests:
            # Extract Library Types
            library_types = [x.library_type.name for x in req.fetched_libraries]
            sample_types = [x.library_type.name for x in req.fetched_samples]

            # Merge the counts
            library_cnt = {
                x[0]: {"libraries": x[1]} for x in Counter(library_types).items()
            }
            sample_cnt = {
                x[0]: {"samples": x[1]} for x in Counter(sample_types).items()
            }
            count = {
                k: {
                    **library_cnt.get(k, {"libraries": 0}),
                    **sample_cnt.get(k, {"samples": 0}),
                }
                for k in library_cnt.keys() | sample_cnt.keys()
            }

            for k, v in count.items():
                temp_dict = counts.get(k, {"libraries": 0, "samples": 0, "total": 0, "requests": 0})
                temp_dict["libraries"] += v["libraries"]
                temp_dict["samples"] += v["samples"]
                temp_dict["total"] += v["libraries"] + v["samples"]
                temp_dict["requests"] += 1
                counts[k] = temp_dict

        return self._get_data(counts)

    def get_pi_counts(self):
        counts = {}

        for req in self.requests:
            pi = req.pi
            pi_name = pi.full_name if pi else "None"
            if pi_name not in counts.keys():
                counts[pi_name] = {"libraries": 0, "samples": 0, "total": 0, "requests": 0}
            counts[pi_name]["libraries"] += len(req.fetched_libraries)
            counts[pi_name]["samples"] += len(req.fetched_samples)
            counts[pi_name]["total"] += len(req.fetched_libraries) + len(req.fetched_samples)
            counts[pi_name]["requests"] += 1

        return self._get_data(counts)

    def get_sequencer_counts(self):
        counts = {}

        for flowcell in self.flowcells:
            sequencer_name = flowcell.pool_size.sequencer.name
            if sequencer_name not in counts.keys():
                counts[sequencer_name] = {"libraries": 0, "samples": 0, "runs": 0}
            counts[sequencer_name]["runs"] += 1

            pools = {x.pool for x in flowcell.fetched_lanes}
            for pool in pools:
                counts[sequencer_name]["libraries"] += len(pool.fetched_libraries)
                counts[sequencer_name]["samples"] += len(pool.fetched_samples)

        requests_counts = Counter(self.requests.values_list(
            'flowcell__pool_size__pool__size__sequencer__name', flat=True))

        data = [
            {
                "name": name,
                "items_count": count["libraries"] + count["samples"],
                "runs_count": count["runs"],
                "requests_count": requests_counts.get(name, None)
            }
            for name, count in counts.items()
            if count["libraries"] + count["samples"] > 0
        ]

        return sorted(data, key=lambda x: x["name"])

    def get_sequencers_list(self):
        return sorted({x.pool_size.sequencer.name for x in self.flowcells})

    def get_pi_sequencer_counts(self):

        # Count requests
        requests = Counter(self.requests.values_list(
            'flowcell__pool_size__pool__size__sequencer__name', 'pi'))
        requests_data = {}
        for item, count in requests.items():
            if item[1] not in requests_data:
                requests_data[item[1]] = {}
            requests_data[item[1]][item[0]] = count
        users = User.objects.filter(id__in=requests_data.keys())
        requests_data = {(str(users.get(id=k)) if k else None): v
                         for k,v in requests_data.items()}

        # Count libraries and samples
        sequencer_mapping = {}
        for flowcell in self.flowcells:  # gets records from flowcell
            sequencer_name = flowcell.pool_size.sequencer.name
            pools = {x.pool for x in flowcell.fetched_lanes}
            for pool in pools:
                records = pool.fetched_libraries + pool.fetched_samples
                for record in records:
                    if record not in sequencer_mapping:
                        sequencer_mapping[record] = []
                    sequencer_mapping[record].append(sequencer_name)
        items = sequencer_mapping.keys()

        pi_mapping = {}
        for req in self.requests:  # gets records from requests
            pi = req.pi
            pi_name = pi.full_name if pi else "None"
            records = req.fetched_libraries + req.fetched_samples

            pi_mapping.update(
                {record: pi_name for record in records if record in items}
            )

        pairs = []
        for k, v in sequencer_mapping.items():
            for sequencer_name in v:
                try:
                    pairs.append((pi_mapping[k], sequencer_name))
                except KeyError:  # KeyError if record exists under flowcell but the correspond. request was deleted.
                    pass  # could add a warning pop-up here
        counts = Counter(pairs)

        data = {}
        for item, count in counts.items():
            if item[0] not in data:
                data[item[0]] = {}
            data[item[0]][f'{item[1]} - Libraries'] = count
            data[item[0]][f'{item[1]} - Requests'] = requests_data.get(item[0], {}).get(item[1], None)

        return OrderedDict(sorted(data.items()))

    def get_sequencing_kit_counts(self):

        # Count requests
        requests_count = Counter(Request.objects.values_list(
            'flowcell__pool_size__pool__size', flat=True))
        pool_sizes = PoolSize.objects.filter(id__in=requests_count.keys())
        requests_count = {str(pool_sizes.get(id=k)): v for
                          k, v in requests_count.items() if k}

        # Count libraries and samples
        counts = {}

        for flowcell in self.flowcells:
            sequencing_kit_name = str(flowcell.pool_size)
            if sequencing_kit_name not in counts.keys():
                counts[sequencing_kit_name] = {"libraries": 0, "samples": 0, "runs": 0}
            counts[sequencing_kit_name]["runs"] += 1

            pools = {x.pool for x in flowcell.fetched_lanes}
            for pool in pools:
                counts[sequencing_kit_name]["libraries"] += len(pool.fetched_libraries)
                counts[sequencing_kit_name]["samples"] += len(pool.fetched_samples)

        data = [
            {
                "name": name,
                "items_count": count["libraries"] + count["samples"],
                "runs_count": count["runs"],
                "requests_count": requests_count.get(name, 0)
            }
            for name, count in counts.items()
            if count["libraries"] + count["samples"] > 0
        ]

        return sorted(data, key=lambda x: x["name"])

    def get_sequencing_kit_list(self):
        return sorted({str(x.pool_size) for x in self.flowcells})

    def get_pi_sequencing_kit_counts(self):

        # Count requests
        requests = Counter(self.requests.values_list(
            'flowcell__pool_size__pool__size', 'pi'))
        requests_data = {}
        pool_size_ids = set()
        for item, count in requests.items():
            pool_size_ids.add(item[0])
            if item[1] not in requests_data:
                requests_data[item[1]] = {}
            requests_data[item[1]][item[0]] = count
        users = User.objects.filter(id__in=requests_data.keys())
        pool_sizes = PoolSize.objects.filter(id__in=pool_size_ids)
        requests_data = {str(users.get(id=k)): {(str(pool_sizes.get(id=l)) if l else None):m
                        for l,m in v.items()}
                        for k,v in requests_data.items()}

        # Count libraries and samples
        sequencer_mapping = {}
        for flowcell in self.flowcells:  # gets records from flowcell
            sequencing_kit_name = str(flowcell.pool_size)
            pools = {x.pool for x in flowcell.fetched_lanes}
            for pool in pools:
                records = pool.fetched_libraries + pool.fetched_samples
                for record in records:
                    if record not in sequencer_mapping:
                        sequencer_mapping[record] = []
                    sequencer_mapping[record].append(sequencing_kit_name)
        items = sequencer_mapping.keys()

        pi_mapping = {}
        for req in self.requests:  # gets records from requests
            pi = req.pi
            pi_name = pi.full_name if pi else "None"
            records = req.fetched_libraries + req.fetched_samples

            pi_mapping.update(
                {record: pi_name for record in records if record in items}
            )

        pairs = []
        for k, v in sequencer_mapping.items():
            for sequencing_kit_name in v:
                try:
                    pairs.append((pi_mapping[k], sequencing_kit_name))
                except KeyError:  # KeyError if record exists under flowcell but the correspond. request was deleted.
                    pass  # could add a warning pop-up here
        counts = Counter(pairs)

        data = {}
        for item, count in counts.items():
            if item[0] not in data:
                data[item[0]] = {}
            data[item[0]][f'{item[1]} - Libraries'] = count
            data[item[0]][f'{item[1]} - Requests'] = requests_data.get(item[0], {}).get(item[1], None)

        return OrderedDict(sorted(data.items()))

    def get_turnaround(self):
        query = """
        CREATE TEMPORARY TABLE IF NOT EXISTS temp1 AS SELECT
            L.id,
            CAST('Library' AS CHAR(7)) rtype,
            L.create_time date1,
            CAST(NULL AS TIMESTAMPTZ) date2
        FROM library_library L;

        CREATE TEMPORARY TABLE IF NOT EXISTS temp2 AS SELECT
            L.id,
            MIN(F.create_time) date3
        FROM library_library L
        LEFT JOIN index_generator_pool_libraries PR
            ON L.id = PR.library_id
        LEFT JOIN index_generator_pool P
            ON PR.pool_id = P.id
        LEFT JOIN flowcell_lane La
            ON P.id = La.pool_id
        LEFT JOIN flowcell_flowcell_lanes FL
            ON La.id = FL.lane_id
        LEFT JOIN flowcell_flowcell F
            ON FL.flowcell_id = F.id
        GROUP BY L.id;

        CREATE TEMPORARY TABLE IF NOT EXISTS temp3 AS SELECT
            S.id,
            CAST('Sample' AS CHAR(6)) rtype,
            S.create_time date1,
            LP.create_time date2
        FROM sample_sample S
        LEFT JOIN library_preparation_librarypreparation LP
            ON S.id = LP.sample_id;

        CREATE TEMPORARY TABLE IF NOT EXISTS temp4 AS SELECT
            S.id,
            MIN(F.create_time) date3
        FROM sample_sample S
        LEFT JOIN index_generator_pool_samples PR
            ON S.id = PR.sample_id
        LEFT JOIN index_generator_pool P
            ON PR.pool_id = P.id
        LEFT JOIN flowcell_lane L
            ON P.id = L.pool_id
        LEFT JOIN flowcell_flowcell_lanes FL
            ON L.id = FL.lane_id
        LEFT JOIN flowcell_flowcell F
            ON FL.flowcell_id = F.id
        GROUP BY S.id;

        SELECT
            rtype, date1, date2, date3
        FROM (
            SELECT *
            FROM temp1 t1
            LEFT JOIN temp2 t2 ON t1.id = t2.id
            UNION ALL
            SELECT *
            FROM temp3 t1
            LEFT JOIN temp4 t2 ON t1.id = t2.id
        ) t
        """

        with connection.cursor() as c:
            c.execute(query)
            columns = [x[0] for x in c.description]
            df = DataFrame(c.fetchall(), columns=columns)

        df["Request -> Preparation"] = df.date2 - df.date1
        df["Preparation -> Sequencing"] = df.date3 - df.date2
        df["Complete Workflow"] = df.date3 - df.date1

        agg_samples_df = df[df.rtype == "Sample"].aggregate(["mean", "std"]).fillna(0)
        agg_samples_df = (agg_samples_df / np.timedelta64(1, "D")).astype(int)

        agg_libraries_df = (
            df[df.rtype == "Library"].iloc[:, -1].aggregate(["mean", "std"]).fillna(0)
        )
        agg_libraries_df = (agg_libraries_df / np.timedelta64(1, "D")).astype(int)

        columns = [
            "Turnaround",
            "Sample (days)",
            "Sample Deviation (days)",
            "Library (days)",
            "Library Deviation (days)",
        ]

        result_df = DataFrame(columns=columns)
        result_df[columns[0]] = df.columns[4:]
        result_df[columns[1]] = agg_samples_df.loc["mean"].values
        result_df[columns[2]] = agg_samples_df.loc["std"].values
        result_df[columns[3]] = [0] * 2 + [agg_libraries_df.loc["mean"]]
        result_df[columns[4]] = [0] * 2 + [agg_libraries_df.loc["std"]]

        return {
            "columns": columns,
            "rows": result_df.T.to_dict().values(),
        }

    @staticmethod
    def _get_data(counts):
        data = [
            {
                "name": name,
                "requests_count": count.get("requests", 0),
                "libraries_count": count["libraries"],
                "samples_count": count["samples"],
                "total_count": count["total"],
            }
            for name, count in counts.items()
            if count["libraries"] + count["samples"] > 0
        ]

        return sorted(data, key=lambda x: x["name"])

def download_report(all_data, start, end):
    """Generate Report as XLSX file"""

    sections = {
        'total_counts': {
            'section_header': 'Total counts',
            'columns': [('type', 'Type'),
                        ('count', 'Total count'),
                        ('count_compromised', 'Count compromised'),
                        ('count_failed', 'Count failed')]
        },
        'organization_counts': {
            'section_header': 'Organization Counts',
            'columns': [('name', 'Organization'),
                        ('requests_count', 'Requests'),
                        ('samples_count', 'Samples'),
                        ('libraries_count', 'Libraries'),
                        ('total_count', 'Samples + Libraries')]
        },
        'protocol_counts': {
            'section_header': 'Protocol Counts',
            'columns': [('name', 'Protocol'),
                        ('requests_count', 'Requests'),
                        ('samples_count', 'Samples'),
                        ('libraries_count', 'Libraries'),
                        ('total_count', 'Samples + Libraries')]
        },
        'library_type_counts': {
            'section_header': 'Library Type Counts',
            'columns': [('name', 'Library Type'),
                        ('requests_count', 'Requests'),
                        ('samples_count', 'Samples'),
                        ('libraries_count', 'Libraries'),
                        ('total_count', 'Samples + Libraries')]
        },
        'pi_counts': {
            'section_header': 'Principal Investigator Counts',
            'columns': [('name', 'Principal Investigator'),
                        ('requests_count', 'Requests'),
                        ('samples_count', 'Samples'),
                        ('libraries_count', 'Libraries'),
                        ('total_count', 'Samples + Libraries')]
        },
        'sequencer_counts': {
            'section_header': 'Sequencer Counts',
            'columns': [('name', 'Sequencer'),
                        ('requests_count', 'Requests'),
                        ('items_count', 'Libraries + Samples'),
                        ('runs_count', 'Runs')]
        },
        'libraries_on_sequencers_counts': {
            'section_header': 'Requests/Libraries on Sequencers',
            'columns': [('pi', 'Principal Investigator')]
        },
        'sequencing_kit_counts': {
            'section_header': 'Sequencing Kit Counts',
            'columns': [('name', 'Sequencing Kit'),
                        ('requests_count', 'Requests'),
                        ('items_count', 'Libraries + Samples'),
                        ('runs_count', 'Runs')]
        },
        'libraries_on_sequencing_kit_counts': {
            'section_header': 'Requests/Libraries on Sequencing Kit',
            'columns': [('pi', 'Principal Investigator')]
        }
    }

    # Create Excel workbook and add as many sheet as there are
    # keys in all_data
    wb = Workbook()
    worksheet_titles = list(all_data.keys())
    [wb.create_sheet('') for _ in worksheet_titles[1:]]

    # Loop through keys of all_data and add each as a sheet
    for ws, (worksheet_title, data) in zip(wb, all_data.items()):

        # Set sheet title
        ws.title = worksheet_title

        # Add filter dates
        row_num = 1
        _cell = ws.cell(row_num, 1, 'Start date')

        # Create styles from default style of first cell
        bold_font = _cell.font.copy(bold=True)
        bold_underline_font = _cell.font.copy(bold=True, underline='single')

        _cell.font = bold_font
        ws.cell(row_num, 2, start.strftime('%d.%m.%Y'))

        row_num += 1
        _cell = ws.cell(row_num, 1, 'End date')
        _cell.font = bold_font
        ws.cell(row_num, 2, end.strftime('%d.%m.%Y'))

        row_num += 2

        # Create sections
        for section_name, section_headers in sections.items():

            section_data = data[section_name]
            section_main_header = section_headers['section_header']
            section_column_names = [c[1] for c in section_headers['columns']]
            section_column_ids = [c[0] for c in section_headers['columns']]

            # Write main section header
            _cell = ws.cell(row_num, 1, section_main_header)
            _cell.font = bold_underline_font
            
            row_num += 2 # Skip one row
            # Write column headers
            
            # Rework section data for libraries_on_sequencers_counts and
            # libraries_on_sequencing_kit_counts
            if section_name == 'libraries_on_sequencers_counts':
                sects = [seq_kit + suffix for seq_kit in data['sequencers_list']
                         for suffix in [' - Requests', ' - Libraries']]
                section_column_names = section_column_names + sects
                section_column_ids = section_column_ids + sects
                section_data = [{**{'pi': pi}, **{s: count.get(s, 0)
                                for s in section_column_ids[1:]}}
                                for pi, count in section_data.items()]

            elif section_name == 'libraries_on_sequencing_kit_counts':
                sects = [seq_kit + suffix for seq_kit in data['sequencing_kit_list']
                         for suffix in [' - Requests', ' - Libraries']]
                section_column_names = section_column_names + sects
                section_column_ids = section_column_ids + sects
                section_data = [{**{'pi': pi}, **{s: count.get(s, 0)
                                for s in section_column_ids[1:]}}
                                for pi, count in section_data.items()]

            for i, column_name in enumerate(section_column_names, 1):
                _cell = ws.cell(row_num, i, column_name)
                _cell.font = bold_font

            # Write data columns
            for d in section_data:
                row_num += 1
                col_num = 1
                for column_id in section_column_ids:
                    ws.cell(row_num, col_num, d[column_id])
                    col_num += 1

            row_num += 3

        for col_idx, _ in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

    return wb


@login_required
# @staff_member_required
def report(request):
    now = timezone.now()
    start = request.GET.get("start", now)
    end = request.GET.get("end", now)

    download = request.GET.get("download", False)
    download = True if download else False

    try:
        start = (
            timezone.datetime.strptime(start, "%d.%m.%Y")
            if type(start) is str
            else start
        )
    except ValueError:
        start = now
    finally:
        start = start.replace(hour=0, minute=0)

    try:
        end = timezone.datetime.strptime(end, "%d.%m.%Y") if type(end) is str else end
    except ValueError:
        end = now
    finally:
        end = end.replace(hour=23, minute=59)

    if start > end:
        start = end.replace(hour=0, minute=0)

    all_data = {}

    # Create report input data for both submitted and sequenced
    # projects
    for status_label, status in [('Submitted', False), ('Sequenced', True)]:

        data = {}
        report = Report(start, end, status)

        # Total Sample Count
        data["total_counts"] = report.get_total_counts()

        # Count by Organization
        data["organization_counts"] = report.get_organization_counts()

        # Count by Library Protocol
        data["protocol_counts"] = report.get_library_protocol_counts()

        # Count by Library Type
        data["library_type_counts"] = report.get_library_type_counts()

        # Count by Principal Investigator
        data["pi_counts"] = report.get_pi_counts()

        # Count by Sequencer
        data["sequencer_counts"] = report.get_sequencer_counts()

        # Count by PI and Sequencer
        data["sequencers_list"] = report.get_sequencers_list()
        data["libraries_on_sequencers_counts"] = report.get_pi_sequencer_counts()

        # Count by Sequencing Kit
        data["sequencing_kit_counts"] = report.get_sequencing_kit_counts()

        # Count by PI and Sequencing Kit
        data["sequencing_kit_list"] = report.get_sequencing_kit_list()
        data["libraries_on_sequencing_kit_counts"] = report.get_pi_sequencing_kit_counts()

        # Count days
        # data["turnaround"] = report.get_turnaround()
        all_data[status_label] = data

    if download:
        wb = download_report(all_data, start, end)

        filename = f"Report_{start.strftime('%d%m%Y')}_{end.strftime('%d%m%Y')}.xlsx"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    return render(request, "report.html", data)


@login_required
@staff_member_required
def database(request):
    return render(request, "database.html")


# @print_sql_queries
@login_required
# @staff_member_required
def database_data(request):
    with connection.cursor() as c:
        query = QUERY.format(
            table_name="library",
            table_name_plural="libraries",
            select=LIBRARY_SELECT,
            joins="",
        )
        c.execute(query)
        columns = [col[0] for col in c.description]
        libraries = [dict(zip(columns, row)) for row in c.fetchall()]

        query = QUERY.format(
            table_name="sample",
            table_name_plural="samples",
            select=SAMPLE_SELECT,
            joins=SAMPLE_JOINS,
        )
        c.execute(query)
        columns = [col[0] for col in c.description]
        samples = [dict(zip(columns, row)) for row in c.fetchall()]

    data = sorted(
        libraries + samples,
        key=lambda x: (
            int(x["Barcode"][:2]),
            int(x["Barcode"][3:]),
        ),
    )

    columns = [
        "Name",
        "Barcode",
        "Status",
        "Request",
        "User",
        "PI",
        "Cost Unit",
        "Organization",
        "Bioinformatician",
        "Library Type",
        "Library Protocol",
        "Concentration (User)",
        "Sample Volume (User)",
        "Sequencing Depth",
        "Read Length",
        "Concentration Method",
        "Equal Representation of Nucleotides",
        "Index Type",
        "Index Reads",
        "Index I7 ID",
        "Index I7",
        "Index I5 ID",
        "Index I5",
        "Amplification Cycles",
        "Dilution Factor",
        "Concentration (Facility)",
        "Sample Volume (Facility)",
        "Amount (Facility)",
        "Size Distribution (Facility)",
        "Concentration Method (Facility)",
        "RNA Quality (Facility)",
        "Organism",
        "Source",
        "Concentration C1",
        "RNA Quality",
        "Nucleic Acid Type",
        "Starting Amount",
        "Spike-in Volume",
        "PCR Cycles",
        "Concentration Library",
        "Mean Fragment Size",
        "nM",
        "qPCR Result",
        "qPCR Result (Facility)",
        "Pool",
        "Flowcell ID",
        "Flowcell create time",
        "Sequencer",
        "Sequencing Kit",
    ]

    return JsonResponse({"columns": columns, "data": data})
